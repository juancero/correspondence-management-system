import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import load_workbook


from services.despacho_service import (
    buscar_pieza,
    pieza_ya_despachada,
    guardar_despacho_completo
)
from services.pdf_service import generar_pdf_despacho


class DespachoScreen(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)

        self.title("Despacho")
        self.geometry("850x600")

        self.id_despacho = None
        self.numero_despacho = None

        self.piezas_temporales = []
        self.codigos_temporales = set()

        self.build_ui()
        self.bloquear_scanner()

    def build_ui(self):
        frame_top = tk.Frame(self)
        frame_top.pack(pady=10)

        self.btn_nuevo = tk.Button(
            frame_top,
            text="Nuevo despacho",
            width=20,
            command=self.nuevo_despacho
        )
        self.btn_nuevo.pack(side=tk.LEFT, padx=5)

        self.btn_cerrar = tk.Button(
            frame_top,
            text="Generar despacho",
            width=20,
            command=self.generar_despacho
        )
        self.btn_cerrar.pack(side=tk.LEFT, padx=5)

        self.btn_cancelar = tk.Button(
            frame_top,
            text="Cancelar despacho",
            width=20,
            command=self.cancelar
        )
        self.btn_cancelar.pack(side=tk.LEFT, padx=5)

        self.btn_importar_excel = tk.Button(
            frame_top,
            text="Importar Excel",
            width=20,
            command=self.importar_excel
        )
        self.btn_importar_excel.pack(side=tk.LEFT, padx=5)

        self.lbl_estado = tk.Label(self, text="Sin despacho activo", fg="red")
        self.lbl_estado.pack(pady=5)

        tk.Label(self, text="Escanear código").pack()

        self.scanner = tk.Entry(self, width=45, font=("Arial", 14))
        self.scanner.pack(pady=8)
        self.scanner.bind("<Return>", self.scan)

        self.lbl_ultima = tk.Label(
            self,
            text="---",
            font=("Arial", 16, "bold")
        )
        self.lbl_ultima.pack(pady=5)

        self.lista = tk.Listbox(self, width=120, height=22)
        self.lista.pack(pady=10)

        self.lbl_total = tk.Label(
            self,
            text="Total: 0",
            font=("Arial", 11, "bold")
        )
        self.lbl_total.pack()

    def nuevo_despacho(self):
        self.id_despacho = None
        self.numero_despacho = "BORRADOR"

        self.piezas_temporales = []
        self.codigos_temporales = set()

        self.lbl_estado.config(
            text="Despacho en preparación",
            fg="orange"
        )

        self.lbl_ultima.config(text="---", fg="black")
        self.lista.delete(0, tk.END)
        self.lbl_total.config(text="Total: 0")

        self.habilitar_scanner()
        self.btn_nuevo.config(state="disabled")

    def scan(self, event):
        codigo = self.scanner.get().strip()
        self.scanner.delete(0, tk.END)

        if self.numero_despacho != "BORRADOR":
            self.lbl_ultima.config(
                text="❌ SIN DESPACHO",
                fg="red"
            )
            return

        if not codigo:
            return

        if codigo in self.codigos_temporales:
            self.lbl_ultima.config(
                text="❌ YA ESCANEADA",
                fg="red"
            )
            return

        pieza = buscar_pieza(codigo)

        if not pieza:
            self.lbl_ultima.config(
                text="❌ NO EXISTE",
                fg="red"
            )
            return

        ya_despachada = pieza_ya_despachada(pieza["id"])

        if ya_despachada:
            self.lbl_ultima.config(
                text=f"❌ YA EN {ya_despachada['numero']}",
                fg="red"
            )
            return

        self.piezas_temporales.append(pieza)
        self.codigos_temporales.add(codigo)

        self.lbl_ultima.config(
            text=f"✔ {pieza['codigo']}",
            fg="green"
        )

        self.refresh_lista()

    def refresh_lista(self):
        self.lista.delete(0, tk.END)

        for p in self.piezas_temporales:
            self.lista.insert(
                tk.END,
                f"{p['codigo']} | {p['apellido_nombre']} | {p['domicilio']} | {p['localidad']}"
            )

        self.lbl_total.config(text=f"Total: {len(self.piezas_temporales)}")

    def importar_excel(self):
        if self.numero_despacho != "BORRADOR":
            messagebox.showwarning(
                "Atención",
                "Primero tenés que iniciar un nuevo despacho."
            )
            return

        ruta = filedialog.askopenfilename(
            title="Seleccionar Excel",
            filetypes=[("Archivos Excel", "*.xlsx")]
        )

        if not ruta:
            return

        try:
            wb = load_workbook(ruta, data_only=True)
            ws = wb.active

            agregadas = 0
            duplicadas = 0
            no_encontradas = 0
            ya_despachadas = 0
            vacias = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                codigo = row[0]

                if codigo is None or str(codigo).strip() == "":
                    vacias += 1
                    continue

                codigo = str(codigo).strip()

                if codigo in self.codigos_temporales:
                    duplicadas += 1
                    continue

                pieza = buscar_pieza(codigo)

                if not pieza:
                    no_encontradas += 1
                    continue

                ya = pieza_ya_despachada(pieza["id"])

                if ya:
                    ya_despachadas += 1
                    continue

                self.piezas_temporales.append(pieza)
                self.codigos_temporales.add(codigo)
                agregadas += 1

            self.refresh_lista()

            self.lbl_ultima.config(
                text=f"✔ Excel importado: {agregadas} agregadas",
                fg="green" if agregadas > 0 else "orange"
            )

            messagebox.showinfo(
                "Importación finalizada",
                f"Resultado de la importación:\n\n"
                f"Agregadas: {agregadas}\n"
                f"Duplicadas en borrador: {duplicadas}\n"
                f"No encontradas: {no_encontradas}\n"
                f"Ya despachadas: {ya_despachadas}\n"
                f"Filas vacías: {vacias}\n\n"
                "Todavía no se guardó nada en la base.\n"
                "Se guardará recién al generar el despacho."
            )

        except Exception as e:
            messagebox.showerror(
                "Error",
                f"No se pudo importar el Excel.\n\n{e}"
            )

    def generar_despacho(self):
        if self.numero_despacho != "BORRADOR":
            messagebox.showwarning(
                "Atención",
                "No hay despacho en preparación."
            )
            return

        if not self.piezas_temporales:
            messagebox.showwarning(
                "Atención",
                "No hay piezas cargadas."
            )
            return

        confirmar = messagebox.askyesno(
            "Generar despacho",
            f"¿Generar despacho?\n\n"
            f"Total de piezas: {len(self.piezas_temporales)}\n\n"
            "Recién ahora se guardará en la base."
        )

        if not confirmar:
            return

        despacho = guardar_despacho_completo(self.piezas_temporales)

        self.id_despacho = despacho["id"]
        self.numero_despacho = despacho["numero"]

        ruta = filedialog.asksaveasfilename(
            title="Guardar PDF del despacho",
            defaultextension=".pdf",
            initialfile=f"despacho_{self.numero_despacho}.pdf",
            filetypes=[("PDF", "*.pdf")]
        )

        if not ruta:
            return

        generar_pdf_despacho(self.id_despacho, ruta)

        messagebox.showinfo(
            "Despacho generado",
            f"Despacho {self.numero_despacho} generado correctamente."
        )

        self.limpiar_pantalla()

    def cancelar(self):
        if self.numero_despacho != "BORRADOR":
            self.limpiar_pantalla()
            return

        if not self.piezas_temporales:
            self.limpiar_pantalla()
            return

        confirmar = messagebox.askyesno(
            "Cancelar despacho",
            "¿Descartar el despacho en preparación?\n\n"
            "No se guardará nada en la base."
        )

        if not confirmar:
            return

        self.limpiar_pantalla()

    def limpiar_pantalla(self):
        self.id_despacho = None
        self.numero_despacho = None

        self.piezas_temporales = []
        self.codigos_temporales = set()

        self.lbl_estado.config(text="Sin despacho activo", fg="red")
        self.lbl_ultima.config(text="---", fg="black")
        self.lista.delete(0, tk.END)
        self.lbl_total.config(text="Total: 0")

        self.bloquear_scanner()
        self.btn_nuevo.config(state="normal")

    def bloquear_scanner(self):
        self.scanner.config(state="disabled")
        self.btn_cerrar.config(state="disabled")
        self.btn_cancelar.config(state="disabled")
        self.btn_importar_excel.config(state="disabled")

    def habilitar_scanner(self):
        self.scanner.config(state="normal")
        self.btn_cerrar.config(state="normal")
        self.btn_cancelar.config(state="normal")
        self.btn_importar_excel.config(state="normal")
        self.scanner.focus_set()
