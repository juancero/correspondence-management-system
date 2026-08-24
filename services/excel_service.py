from openpyxl import Workbook
from openpyxl.styles import Font
from database import get_connection


def generar_excel_rendicion(id_rendicion, ruta):
    with get_connection() as conn:
        cur = conn.cursor()

        cur.execute("""
            SELECT numero, fecha
            FROM rendiciones
            WHERE id = ?
        """, (id_rendicion,))
        rendicion = cur.fetchone()

        cur.execute("""
            SELECT 
                c.codigo,
                c.apellido_nombre,
                c.domicilio,
                c.localidad,
                rd.resultado,
                rd.fecha_rendicion
            FROM rendicion_detalle rd
            JOIN correspondencias c ON c.id = rd.id_correspondencia
            WHERE rd.id_rendicion = ?
            ORDER BY rd.id ASC
        """, (id_rendicion,))
        filas = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Rendición"

    ws["A1"] = f"Rendición N° {rendicion['numero']}"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = f"Fecha: {rendicion['fecha']}"
    ws["A3"] = f"Total piezas rendidas: {len(filas)}"

    headers = [
        "Código",
        "Apellido y Nombre",
        "Domicilio",
        "Localidad",
        "Resultado",
        "Fecha Rendición",
        "Nro Rendición"
    ]

    ws.append([])
    ws.append(headers)

    header_row = 5

    for col in range(1, len(headers) + 1):
        ws.cell(row=header_row, column=col).font = Font(bold=True)

    for f in filas:
        ws.append([
            f["codigo"],
            f["apellido_nombre"],
            f["domicilio"],
            f["localidad"],
            f["resultado"],
            f["fecha_rendicion"],
            rendicion["numero"]
        ])

    widths = {
        "A": 18,
        "B": 35,
        "C": 40,
        "D": 25,
        "E": 25,
        "F": 18,
        "G": 18,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(ruta)


def generar_excel_estado_despacho(id_despacho, ruta):
    with get_connection() as conn:
        cur = conn.cursor()

        cur.execute("""
            SELECT numero, fecha, estado
            FROM despachos
            WHERE id = ?
        """, (id_despacho,))
        despacho = cur.fetchone()

        if not despacho:
            raise ValueError("Despacho no encontrado")

        cur.execute("""
            SELECT 
                c.codigo,
                c.apellido_nombre,
                c.domicilio,
                c.localidad,
                c.estado,
                c.resultado_rendicion,
                c.fecha_rendicion,
                c.numero_rendicion
            FROM despacho_detalle dd
            JOIN correspondencias c ON c.id = dd.id_correspondencia
            WHERE dd.id_despacho = ?
            ORDER BY dd.id ASC
        """, (id_despacho,))
        filas = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Estado Despacho"

    ws["A1"] = f"Estado Despacho N° {despacho['numero']}"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = f"Fecha: {despacho['fecha']}"
    ws["A3"] = f"Estado despacho: {despacho['estado']}"
    ws["A4"] = f"Total piezas: {len(filas)}"

    headers = [
        "Código",
        "Apellido y Nombre",
        "Domicilio",
        "Localidad",
        "Estado Pieza",
        "Resultado Rendición",
        "Fecha Rendición",
        "Nro Rendición"
    ]

    ws.append([])
    ws.append(headers)

    header_row = 6

    for col in range(1, len(headers) + 1):
        ws.cell(row=header_row, column=col).font = Font(bold=True)

    for f in filas:
        ws.append([
            f["codigo"],
            f["apellido_nombre"],
            f["domicilio"],
            f["localidad"],
            f["estado"],
            f["resultado_rendicion"] or "",
            f["fecha_rendicion"] or "",
            f["numero_rendicion"] or "",
        ])

    widths = {
        "A": 18,
        "B": 35,
        "C": 40,
        "D": 25,
        "E": 18,
        "F": 25,
        "G": 18,
        "H": 18,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(ruta)
