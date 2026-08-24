import csv
import sqlite3
from pathlib import Path
from openpyxl import load_workbook

from database import get_connection


COLUMNAS_VALIDAS = {
    "codigo",
    "destinatario",
    "direccion",
    "localidad",
    "provincia",
    "observacion",
}


def normalizar_header(valor):
    if valor is None:
        return ""
    return str(valor).strip().lower().replace(" ", "_")


def importar_archivo(ruta_archivo):
    ruta = Path(ruta_archivo)

    if not ruta.exists():
        raise FileNotFoundError("El archivo seleccionado no existe.")

    if ruta.suffix.lower() == ".csv":
        filas = leer_csv(ruta)
    elif ruta.suffix.lower() in [".xlsx", ".xlsm"]:
        filas = leer_excel(ruta)
    else:
        raise ValueError("Formato no soportado. Usá CSV o XLSX.")

    if not filas:
        return {
            "insertados": 0,
            "duplicados": 0,
            "errores": ["El archivo está vacío."]
        }

    return insertar_correspondencias(filas)


def leer_csv(ruta):
    filas = []

    with open(ruta, "r", encoding="utf-8-sig", newline="") as archivo:
        reader = csv.DictReader(archivo)
        for row in reader:
            filas.append({
                normalizar_header(k): v
                for k, v in row.items()
            })

    return filas


def leer_excel(ruta):
    filas = []

    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb.active

    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(normalizar_header(cell.value))

    for row in ws.iter_rows(min_row=2, values_only=True):
        fila = {}
        for index, value in enumerate(row):
            if index < len(headers):
                fila[headers[index]] = value
        filas.append(fila)

    wb.close()
    return filas


def insertar_correspondencias(filas):
    insertados = 0
    duplicados = 0
    errores = []

    codigos_en_archivo = set()

    with get_connection() as conn:
        cursor = conn.cursor()

        for nro_fila, fila in enumerate(filas, start=2):
            codigo = str(fila.get("cod_barra", "")).strip()

            if not codigo:
                errores.append(f"Fila {nro_fila}: código vacío.")
                continue

            if codigo in codigos_en_archivo:
                duplicados += 1
                errores.append(f"Fila {nro_fila}: código duplicado en archivo: {codigo}")
                continue

            codigos_en_archivo.add(codigo)

            datos = {
                "codigo": codigo,
                "nro_envio_orden": str(fila.get("nro_envio/orden", "") or "").strip(),
                "nro_afiliado": str(fila.get("nro_afiliado", "") or "").strip(),
                "apellido_nombre": str(fila.get("apellido_y_nombre", "") or "").strip(),
                "domicilio": str(fila.get("domicilio", "") or "").strip(),
                "cp": str(fila.get("cp", "") or "").strip(),
                "localidad": str(fila.get("localidad", "") or "").strip(),
                "distrito_electoral": str(fila.get("districto_electoral", "") or "").strip(),
                "observacion": str(fila.get("observacion", "") or "").strip(),
            }

            try:
                cursor.execute("""
                    INSERT INTO correspondencias (
                        codigo,
                        nro_envio_orden,
                        nro_afiliado,
                        apellido_nombre,
                        domicilio,
                        cp,
                        localidad,
                        distrito_electoral,
                        observacion
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    datos["codigo"],
                    datos["nro_envio_orden"],
                    datos["nro_afiliado"],
                    datos["apellido_nombre"],
                    datos["domicilio"],
                    datos["cp"],
                    datos["localidad"],
                    datos["distrito_electoral"],
                    datos["observacion"],
                ))

                insertados += 1

            except sqlite3.IntegrityError:
                duplicados += 1
                errores.append(f"Fila {nro_fila}: código ya existe en la base: {codigo}")

        conn.commit()

    return {
        "insertados": insertados,
        "duplicados": duplicados,
        "errores": errores
    }